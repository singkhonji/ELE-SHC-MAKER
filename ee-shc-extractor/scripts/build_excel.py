"""
build_excel.py — Step 5: Build formatted Load Schedule Excel workbook

Usage (when imported):
    build_excel(boards_data, output_path) -> str   (returns output path)

boards_data: list of dicts from parse_board()
output_path: str path for .xlsx output
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colour palette ───────────────────────────────────────────────────────────
DARK_BLUE  = '1F3864'
MID_BLUE   = '2E75B6'
LT_BLUE    = 'D6E4F0'
ORANGE_BG  = 'FCE4D6'
SPARE_BG   = 'E2EFDA'
YELLOW_BG  = 'FFF2CC'
GREY_BG    = 'F2F2F2'
WHITE      = 'FFFFFF'
TAB_COLORS = [DARK_BLUE, MID_BLUE, '70AD47', 'ED7D31']  # cycle for >2 boards

# ─── Column widths ────────────────────────────────────────────────────────────
COL_WIDTHS = [11, 36, 10, 14, 9, 10, 14, 11, 12, 10, 32, 22, 22, 9]
HDR_LABELS = [
    'Circuit No.', 'Description / Load Name', 'CB Rating (A)', 'CB Type',
    'No. of Poles', 'Breaking Cap.', 'Connected Load (kW)', 'Demand Factor',
    'Max Demand (kW)', 'Current (A)', 'Cable Size', 'Cable Route',
    'Remarks', 'Status',
]

# ─── Style helpers ────────────────────────────────────────────────────────────
_thin = Side(style='thin', color='BFBFBF')
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _cs(ws, r, c, v='', bold=False, italic=False, fc='000000',
        bg=None, ha='center', wrap=False, sz=9):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name='Arial', bold=bold, italic=italic, color=fc, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border    = _bdr
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    return cell


def _ms(ws, r1, c1, r2, c2, v='', bold=False, fc='000000',
        bg=None, ha='center', wrap=False, sz=9):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=v)
    cell.font      = Font(name='Arial', bold=bold, color=fc, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border    = _bdr
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    return cell


def _col_headers(ws, row):
    ws.row_dimensions[row].height = 28
    for ci, h in enumerate(HDR_LABELS, 1):
        _cs(ws, row, ci, h, bold=True, bg=DARK_BLUE, fc=WHITE, wrap=True)


# ─── Sheet builder ────────────────────────────────────────────────────────────
def _build_sheet(ws, bdata: dict, tab_color: str):
    ws.sheet_properties.tabColor = tab_color

    # Col widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bid          = bdata.get('id') or 'UNKNOWN'
    supply       = bdata.get('supply') or '—'
    incomer_cb   = bdata.get('incomer_cb') or '— (refer drawing)'
    incomer_cable= bdata.get('incomer_cable') or '—'
    busbar       = bdata.get('busbar') or '—'
    cable_dc     = bdata.get('cable_dc', False)
    dc_warn      = ' ⚠️ CABLE BY DC MEP' if cable_dc else ''

    # Parse incomer A rating & breaking capacity
    inc_a_m  = re.search(r'(\d+A)', incomer_cb)
    inc_a    = inc_a_m.group(1) if inc_a_m else '—'
    inc_brk  = '50kA' if '50kA' in incomer_cb else '36kA'

    # ── Row 1-2: Title ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    _ms(ws, 1, 1, 1, 14, f'LOAD SCHEDULE — {bid}',
        bold=True, fc=WHITE, bg=DARK_BLUE, sz=13)
    _ms(ws, 2, 1, 2, 14, 'DayOne CTP Building D  |  Common Area  |  Electrical Load Schedule',
        fc=WHITE, bg=MID_BLUE, sz=9)

    # ── Row 3-5: Board info ───────────────────────────────────────────────
    info_rows = [
        ('Board Designation:', bid,                        'Supply Source:',  supply),
        ('Location:',          'Common Area (1 NO.)',      'Incomer CB:',     incomer_cb),
        ('Busbar Rating:',     busbar,                     'Incomer Cable:',  incomer_cable + dc_warn),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows):
        r = 3 + i
        ws.row_dimensions[r].height = 16
        _cs(ws, r, 1, l1, bold=True, bg=LT_BLUE, ha='left')
        _ms(ws, r, 2, r, 5, v1, bg=WHITE, ha='left', wrap=True)
        _cs(ws, r, 6, l2, bold=True, bg=LT_BLUE, ha='left')
        _ms(ws, r, 7, r, 14, v2, bg=WHITE, ha='left', wrap=True)

    # Separator
    ws.row_dimensions[6].height = 5
    for c in range(1, 15):
        ws.cell(row=6, column=c).fill = PatternFill('solid', fgColor=MID_BLUE)

    # ── Section A: Incomer ────────────────────────────────────────────────
    ws.row_dimensions[7].height = 16
    _ms(ws, 7, 1, 7, 14, 'SECTION A — INCOMER', bold=True, fc=WHITE, bg=MID_BLUE)
    _col_headers(ws, 8)

    ws.row_dimensions[9].height = 22
    inc_remark = incomer_cb + ('  ⚠️ CABLE BY DC MEP' if cable_dc else '')
    inc_vals = [
        'INCOMER', f'Supply from {supply}',
        inc_a, 'MCCB (L,S,I)', '3P+N', inc_brk,
        '—', '—', '—', '—',
        incomer_cable, 'Cable Ladder / Tray', inc_remark, 'ACTIVE',
    ]
    for ci, v in enumerate(inc_vals, 1):
        _cs(ws, 9, ci, v, bold=(ci <= 2), bg=ORANGE_BG,
            ha='left' if ci in (2, 11, 12, 13) else 'center', wrap=True)

    # ── Section B: Outgoing circuits ──────────────────────────────────────
    ws.row_dimensions[10].height = 16
    _ms(ws, 10, 1, 10, 14, 'SECTION B — OUTGOING CIRCUITS (FEEDERS TO SUB-BOARDS)',
        bold=True, fc=WHITE, bg=MID_BLUE)
    _col_headers(ws, 11)

    circuits = bdata.get('circuits', {})
    for i, (ckt, d) in enumerate(circuits.items()):
        r = 12 + i
        ws.row_dimensions[r].height = 26
        spare = d.get('spare', False)
        bg    = SPARE_BG if spare else (LT_BLUE if i % 2 == 0 else WHITE)
        fc_t  = '7F7F7F' if spare else '000000'
        row_v = [
            ckt, d['desc'], d['cb_a'], 'MCCB', '3P+N', d['cb_brk'],
            '', '', '', '',
            d['cable'], d['route'], '', d['status'],
        ]
        for ci, v in enumerate(row_v, 1):
            _cs(ws, r, ci, v, bold=(ci == 1), italic=spare, bg=bg, fc=fc_t,
                ha='left' if ci in (2, 11, 12) else 'center', wrap=True)

    # ── Section C: Accessories ────────────────────────────────────────────
    ac_start = 12 + len(circuits)
    ws.row_dimensions[ac_start].height = 16
    _ms(ws, ac_start, 1, ac_start, 14,
        'SECTION C — PANEL ACCESSORIES & METERING',
        bold=True, fc=WHITE, bg=MID_BLUE)

    ws.row_dimensions[ac_start + 1].height = 18
    acc_hdr = ['Item', 'Description', 'Standard / Spec', 'Qty', 'Remarks',
               '', '', '', '', '', '', '', '', '']
    for ci, h in enumerate(acc_hdr, 1):
        _cs(ws, ac_start + 1, ci, h, bold=bool(h), bg=DARK_BLUE, fc=WHITE)

    for j, acc in enumerate(bdata.get('acc', [])):
        r  = ac_start + 2 + j
        bg = LT_BLUE if j % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 18
        _cs(ws, r, 1, acc[0], bg=bg)
        _cs(ws, r, 2, acc[1], bg=bg, ha='left')
        _cs(ws, r, 3, acc[2], bg=bg, ha='left')
        _cs(ws, r, 4, acc[3], bg=bg)
        _ms(ws, r, 5, r, 14, acc[4], bg=bg, ha='left')

    # ── Notes ─────────────────────────────────────────────────────────────
    n_start = ac_start + 2 + len(bdata.get('acc', []))
    ws.row_dimensions[n_start].height = 16
    _ms(ws, n_start, 1, n_start, 14, 'NOTES', bold=True, fc=WHITE, bg=MID_BLUE)

    notes = [
        f'1.  Board: {bid} — Common Area (1 NO.)',
        f'2.  Incomer supply from {supply} | {incomer_cb}.',
        f'3.  Incomer cable: {incomer_cable}.',
    ]
    note_n = 4
    if cable_dc:
        notes.append(
            f'{note_n}.  ⚠️  SCOPE CONFLICT: "CABLE BY DC MEP" annotated — '
            'confirm GC vs DC MEP boundary before pricing.'
        )
        note_n += 1

    acc_labels = [a[1] for a in bdata.get('acc', [])]
    if not any('ELR' in a for a in acc_labels):
        notes.append(
            f'{note_n}.  ⚠️  ELR not found as standalone text — verify with M&E Engineer.'
        )
        note_n += 1
    if not any('ZCT' in a for a in acc_labels):
        notes.append(
            f'{note_n}.  ⚠️  ZCT not found as standalone text — verify with M&E Engineer.'
        )
        note_n += 1

    notes += [
        f'{note_n}.  Installed capacity (kW) and demand factor to be populated by M&E Engineer.',
        f'{note_n+1}.  Drawing source: DWG Data Extraction XLS. Verify against IFC drawings before pricing.',
    ]

    for k, note in enumerate(notes):
        r  = n_start + 1 + k
        bg = YELLOW_BG if '⚠️' in note else (GREY_BG if k % 2 == 0 else WHITE)
        ws.row_dimensions[r].height = 18
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        cell = ws.cell(row=r, column=1, value=note)
        cell.font      = Font(name='Arial', size=9, italic='⚠️' in note)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border    = _bdr

    # ── Freeze & page setup ───────────────────────────────────────────────
    ws.freeze_panes = 'A12'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ─── Main entry point ─────────────────────────────────────────────────────────
def build_excel(boards_data: list[dict], output_path: str) -> str:
    wb = Workbook()

    for i, bdata in enumerate(boards_data):
        bid = bdata.get('id') or f'Board_{i+1}'
        # Sanitise sheet name (Excel limit: 31 chars, no special chars)
        sheet_name = re.sub(r'[\\/*?:\[\]]', '-', bid)[:31]
        tab_color  = TAB_COLORS[i % len(TAB_COLORS)]

        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        _build_sheet(ws, bdata, tab_color)

    wb.save(output_path)
    return output_path


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import sys, os, json
    if len(sys.argv) < 3:
        print("Usage: python scripts/build_excel.py <boards_data.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        data = json.load(f)
    out = build_excel(data, sys.argv[2])
    print(f"Saved: {out}")
