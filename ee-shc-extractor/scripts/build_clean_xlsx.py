"""
build_clean_xlsx.py — Write the "clean" review workbook from parsed board data.

Single worksheet, one per-circuit table per cabinet, with a labelled header
block separating each cabinet.  The layout is **round-trippable**: read_clean_xlsx.py
parses it back into board dicts, so the user can review/edit this file and feed it
to Stage 2 (Load Schedule).

Usage (when imported):
    build_clean_xlsx(boards_data, output_path) -> str   (returns output path)

boards_data: list of dicts from parse_board()
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colour palette (shared look with build_excel.py) ──────────────────────────
DARK_BLUE = '1F3864'
MID_BLUE  = '2E75B6'
LT_BLUE   = 'D6E4F0'
SPARE_BG  = 'E2EFDA'
WARN_BG   = 'FFF2CC'
WHITE     = 'FFFFFF'

# ─── Stable markers / labels (imported by read_clean_xlsx.py — keep in sync) ────
CABINET_MARKER = 'CABINET:'
LBL_SUPPLY   = 'Supply Source:'
LBL_INC_CB   = 'Incomer CB:'
LBL_INC_CAB  = 'Incomer Cable:'
LBL_BUSBAR   = 'Busbar:'
LBL_ACC      = 'Accessories:'
LBL_FLAGS    = 'Flags:'
HEADER_LABELS = [LBL_SUPPLY, LBL_INC_CB, LBL_INC_CAB, LBL_BUSBAR, LBL_ACC, LBL_FLAGS]

# ─── Circuit table columns (fixed positions — reader maps by index) ─────────────
CIRCUIT_COLS = [
    'Circuit No.', 'Description / Load Name', 'CB Rating (A)', 'CB Type',
    'Poles', 'Breaking Cap.', 'Connected Load (kW)', 'Demand Factor',
    'Cable Size', 'Cable Route', 'Status', 'Remarks',
]
COL_WIDTHS = [11, 34, 12, 10, 7, 11, 16, 13, 38, 20, 9, 24]
NCOLS = len(CIRCUIT_COLS)

_thin = Side(style='thin', color='BFBFBF')
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _cell(ws, r, c, v='', bold=False, italic=False, fc='000000',
          bg=None, ha='center', wrap=False, sz=9):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name='Arial', bold=bold, italic=italic, color=fc, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border    = _bdr
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    return cell


def _merge(ws, r, c1, c2, v='', bold=False, italic=False, fc='000000', bg=None,
           ha='left', wrap=False, sz=9):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = _cell(ws, r, c1, v, bold=bold, italic=italic, fc=fc, bg=bg,
                 ha=ha, wrap=wrap, sz=sz)
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = _bdr
    return cell


def _acc_summary(bdata) -> str:
    """Compact accessory tag list, e.g. 'SPD, CT 400/5A, PM, BMS, ELR, ZCT'."""
    tags = []
    if bdata.get('spd'):
        tags.append('SPD')
    if bdata.get('ct'):
        tags.append(f"CT {bdata['ct']}")
    labels = ' '.join(a[1] for a in bdata.get('acc', []))
    for key, tag in (('Power Meter', 'PM'), ('BMS', 'BMS'),
                     ('Earth Leakage', 'ELR'), ('Zero Core', 'ZCT')):
        if key in labels:
            tags.append(tag)
    # ' | ' separator (not ',') — CT specs like "400/5A 5P20,20VA" contain commas
    return ' | '.join(tags) if tags else '—'


def _header_row(ws, row, label, value, value_bg=WHITE):
    _merge(ws, row, 1, 2, label, bold=True, bg=LT_BLUE, ha='left')
    warn = '⚠️' in str(value) or str(value).strip() in ('', '—')
    _merge(ws, row, 3, NCOLS, value, bg=(WARN_BG if warn else value_bg),
           ha='left', wrap=True)


def _write_board(ws, row: int, bdata: dict) -> int:
    bid     = bdata.get('id') or 'UNKNOWN'
    supply  = bdata.get('supply') or '—'
    incomer = bdata.get('incomer_cb') or '⚠️ — refer drawing'
    inc_cab = bdata.get('incomer_cable') or '⚠️ — not in DXF, verify'
    busbar  = bdata.get('busbar') or '—'
    flags   = '⚠️ CABLE BY DC MEP' if bdata.get('cable_dc') else '—'

    # Cabinet title bar (stable marker for the reader)
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, NCOLS, f'{CABINET_MARKER} {bid}',
           bold=True, fc=WHITE, bg=DARK_BLUE, sz=12)
    row += 1

    # Labelled header block
    for lbl, val in (
        (LBL_SUPPLY,  supply),
        (LBL_INC_CB,  incomer),
        (LBL_INC_CAB, inc_cab),
        (LBL_BUSBAR,  busbar),
        (LBL_ACC,     _acc_summary(bdata)),
        (LBL_FLAGS,   flags),
    ):
        ws.row_dimensions[row].height = 16
        _header_row(ws, row, lbl, val)
        row += 1

    # Circuit table header
    ws.row_dimensions[row].height = 28
    for c, h in enumerate(CIRCUIT_COLS, 1):
        _cell(ws, row, c, h, bold=True, bg=MID_BLUE, fc=WHITE, wrap=True)
    row += 1

    # Circuit rows
    for i, (ckt, d) in enumerate(bdata.get('circuits', {}).items()):
        spare = d.get('spare', False)
        bg    = SPARE_BG if spare else (LT_BLUE if i % 2 == 0 else WHITE)
        fc_t  = '7F7F7F' if spare else '000000'
        vals = [
            ckt, d.get('desc', ''), d.get('cb_a', '--'), 'MCCB', '3P+N',
            d.get('cb_brk', '36kA'),
            d.get('conn_load', ''),          # Connected Load (kW) — user fills
            d.get('demand', ''),             # Demand Factor — user fills
            d.get('cable', '--'), d.get('route', '--'),
            d.get('status', ''),
            'Spare Way' if spare else d.get('remark', ''),
        ]
        ws.row_dimensions[row].height = 26
        for c, v in enumerate(vals, 1):
            _cell(ws, row, c, v, bold=(c == 1), italic=spare, bg=bg, fc=fc_t,
                  ha='left' if c in (2, 9, 10, 12) else 'center', wrap=True)
        row += 1

    # Blank spacer (separates cabinets — also the reader's end-of-table marker)
    row += 1
    return row


def build_clean_xlsx(boards_data: list, output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clean Summary'

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    for bdata in boards_data:
        row = _write_board(ws, row, bdata)

    ws.freeze_panes = 'A1'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1

    wb.save(output_path)
    return output_path


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import sys, json
    if len(sys.argv) < 3:
        print("Usage: python scripts/build_clean_xlsx.py <boards_data.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding='utf-8') as f:
        data = json.load(f)
    print("Saved:", build_clean_xlsx(data, sys.argv[2]))
