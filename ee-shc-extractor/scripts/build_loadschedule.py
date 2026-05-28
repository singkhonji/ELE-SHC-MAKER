"""
build_loadschedule.py — Stage 2 workbook builder.

Builds the final Load Schedule workbook from board dicts (read_clean_xlsx):
    • INDEX sheet (first) — clickable list of every cabinet
    • DIAGRAM sheet — clickable box tree of the supply hierarchy
    • one Load Schedule sheet per cabinet (reuses build_excel._build_sheet)

Usage (when imported):
    build_loadschedule(boards, output_path) -> str
"""

import os
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_excel import _build_sheet, sheet_name, DARK_BLUE, MID_BLUE, LT_BLUE, WHITE

GREEN     = '70AD47'
GREY      = 'BFBFBF'
PALE      = 'F2F2F2'
LINK_BLUE = '0563C1'

INDEX_SHEET   = 'INDEX'
DIAGRAM_SHEET = 'DIAGRAM'
BOARD_RE      = re.compile(r'^D\d+-(SB|MDB|DB|EMSB)-', re.IGNORECASE)

_thin = Side(style='thin', color=GREY)
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _board_type(bid: str) -> str:
    m = re.search(r'-(EMSB|MDB|SB|DB)-', bid or '', re.IGNORECASE)
    return m.group(1).upper() if m else '—'


def _link(cell, target_sheet: str):
    # Internal navigation must use `location` (not an external target) so Excel
    # treats it as a same-workbook jump.
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{sheet_name(target_sheet)}'!A1",
        display=str(cell.value or ''),
    )


# ─── INDEX ──────────────────────────────────────────────────────────────────
def _build_index(ws, boards):
    headers = ['#', 'Board ID', 'Type', 'Supply From',
               'Circuits', 'Active', 'Spare', 'Incomer CB', 'Flags']
    widths  = [5, 26, 7, 24, 9, 8, 8, 36, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:I1')
    c = ws.cell(1, 1, 'LOAD SCHEDULE — INDEX')
    c.font = Font(name='Arial', bold=True, color=WHITE, size=14)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:I2')
    c = ws.cell(2, 1, 'DayOne CTP Building D | Common Area | คลิกชื่อตู้เพื่อไปยังหน้าตาราง • ดูผังการจ่ายไฟที่ DIAGRAM')
    c.font = Font(name='Arial', color=WHITE, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=MID_BLUE)
    _link(ws.cell(2, 1), DIAGRAM_SHEET)

    hr = 3
    for i, h in enumerate(headers, 1):
        cell = ws.cell(hr, i, h)
        cell.font = Font(name='Arial', bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor=DARK_BLUE)
        cell.border = _bdr
    ws.row_dimensions[hr].height = 22

    for idx, b in enumerate(boards, 1):
        r = hr + idx
        circuits = b.get('circuits', {})
        n_act = sum(1 for d in circuits.values() if not d.get('spare'))
        n_spr = sum(1 for d in circuits.values() if d.get('spare'))
        flags = '⚠️ CABLE BY DC MEP' if b.get('cable_dc') else ''
        vals = [idx, b.get('id', ''), _board_type(b.get('id', '')),
                b.get('supply', '—'), len(circuits), n_act, n_spr,
                b.get('incomer_cb', '—'), flags]
        bg = LT_BLUE if idx % 2 else WHITE
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.alignment = Alignment(
                horizontal='left' if ci in (2, 4, 8, 9) else 'center',
                vertical='center', wrap_text=True)
            cell.border = _bdr
            cell.fill = PatternFill('solid', fgColor='FFF2CC' if (ci == 9 and flags) else bg)
            cell.font = Font(name='Arial', size=9)
        # Board ID → hyperlink to its sheet
        link = ws.cell(r, 2)
        _link(link, b.get('id', ''))
        link.font = Font(name='Arial', size=9, bold=True, color=LINK_BLUE, underline='single')
        ws.row_dimensions[r].height = 24

    ws.freeze_panes = 'A4'
    ws.sheet_view.showGridLines = False


# ─── DIAGRAM (clickable box tree) ─────────────────────────────────────────────
def _hierarchy(boards):
    present = {b['id'] for b in boards if b.get('id')}
    children = defaultdict(list)
    seen_edge = set()
    parent_of = {}

    def add_edge(p, c):
        if p and c and (p, c) not in seen_edge:
            seen_edge.add((p, c))
            children[p].append(c)
            parent_of[c] = p

    for b in boards:
        bid = b.get('id')
        if b.get('supply'):
            add_edge(b['supply'], bid)
        for d in b.get('circuits', {}).values():
            desc = (d.get('desc') or '').strip()
            if BOARD_RE.match(desc) and desc != bid:
                add_edge(bid, desc)

    all_nodes = set(present) | set(parent_of) | set(children)
    roots = sorted(n for n in all_nodes if n not in parent_of)

    # Levels via BFS
    level = {}
    order = []
    for root in roots:
        level[root] = 0
        queue = [root]
        while queue:
            n = queue.pop(0)
            order.append(n)
            for ch in children.get(n, []):
                if ch not in level:
                    level[ch] = level[n] + 1
                    queue.append(ch)
    for n in all_nodes:
        level.setdefault(n, 0)

    # Slot (x) assignment: leaves sequential, parents centred on children
    slot = {}
    counter = [0]

    def assign(n):
        kids = children.get(n, [])
        if not kids:
            slot[n] = counter[0]
            counter[0] += 1
        else:
            for k in kids:
                assign(k)
            slot[n] = sum(slot[k] for k in kids) / len(kids)

    for root in roots:
        assign(root)

    return present, children, roots, level, slot


def _build_diagram(ws, boards):
    present, children, roots, level, slot = _hierarchy(boards)
    if not slot:
        return

    BOX_W, GAP, BOX_H, CONN = 6, 2, 3, 2
    PITCH = BOX_W + GAP
    LANE  = BOX_H + CONN
    TOP   = 4   # rows reserved for the title

    max_slot = max(slot.values())
    n_cols = int(round(max_slot)) * PITCH + BOX_W + 4
    for ci in range(1, n_cols + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 3.0

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(n_cols, 30))
    t = ws.cell(1, 1, 'แผนผังการจ่ายไฟ (POWER DISTRIBUTION DIAGRAM) — คลิกกล่องตู้เพื่อไปหน้าตาราง')
    t.font = Font(name='Arial', bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal='left', vertical='center')
    t.fill = PatternFill('solid', fgColor=DARK_BLUE)
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols, 30))
    bk = ws.cell(2, 1, '⟵ กลับไปหน้า INDEX')
    bk.font = Font(name='Arial', color=LINK_BLUE, size=9, underline='single')
    _link(bk, INDEX_SHEET)

    pos = {}  # node -> (row0, col0, center_col)

    def node_rc(n):
        col0 = 1 + int(round(slot[n] * PITCH))
        row0 = TOP + level[n] * LANE
        return row0, col0

    # Draw boxes
    for n in slot:
        row0, col0 = node_rc(n)
        col1 = col0 + BOX_W - 1
        row1 = row0 + BOX_H - 1
        center = col0 + BOX_W // 2
        pos[n] = (row0, col0, center)

        ws.merge_cells(start_row=row0, start_column=col0, end_row=row1, end_column=col1)
        cell = ws.cell(row0, col0)
        is_present = n in present
        typ = _board_type(n)
        fill = {'EMSB': GREY, 'SB': DARK_BLUE, 'MDB': MID_BLUE, 'DB': GREEN}.get(typ, PALE)
        if not is_present:
            fill = PALE
        label = n if is_present else f'{n}\n(ref)'
        cell.value = label
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor=fill)
        fc = WHITE if (is_present and typ in ('SB', 'MDB')) else ('000000' if not is_present else WHITE)
        if is_present and typ == 'DB':
            fc = WHITE
        cell.font = Font(name='Arial', bold=True, size=9,
                         color=(fc if is_present else '7F7F7F'),
                         underline=('single' if is_present else None))
        side = Side(style='medium', color=('1F3864' if is_present else GREY))
        bdr = Border(left=side, right=side, top=side, bottom=side)
        for rr in range(row0, row1 + 1):
            for cc in range(col0, col1 + 1):
                ws.cell(rr, cc).border = bdr
        if is_present:
            _link(cell, n)

    # Draw connectors (box-drawing chars, one char per cell)
    def put(r, c, ch):
        cell = ws.cell(r, c)
        cell.value = ch
        cell.font = Font(name='Consolas', size=9, color='404040')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for parent, kids in children.items():
        if parent not in pos or not kids:
            continue
        prow0, pcol0, pcenter = pos[parent]
        p_bottom = prow0 + BOX_H - 1
        kid_centers = sorted(pos[k][2] for k in kids if k in pos)
        if not kid_centers:
            continue
        bus_row   = p_bottom + 1
        child_top = prow0 + LANE
        lo, hi    = min(kid_centers), max(kid_centers)

        # Single child aligned under parent → straight vertical, no bus.
        if len(kid_centers) == 1 and kid_centers[0] == pcenter:
            for rr in range(bus_row, child_top):
                put(rr, pcenter, '│')
            continue

        span_lo, span_hi = min(lo, pcenter), max(hi, pcenter)
        # horizontal bus (plain line first)
        for cc in range(span_lo, span_hi + 1):
            put(bus_row, cc, '─')
        # child branch junctions
        for cc in kid_centers:
            put(bus_row, cc, '┬')
        # parent junction (joins the box directly above)
        if lo <= pcenter <= hi:
            put(bus_row, pcenter, '┴' if pcenter not in kid_centers else '┼')
        elif pcenter < lo:
            put(bus_row, pcenter, '┌')
        else:
            put(bus_row, pcenter, '┐')
        # verticals from bus down to each child top
        for cc in kid_centers:
            for rr in range(bus_row + 1, child_top):
                put(rr, cc, '│')

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'


# ─── Per-board accessory reconstruction (standard boilerplate specs) ───────────
def _reconstruct_acc(b):
    tags = b.get('acc_tags', [])
    ct   = b.get('ct', '')
    busbar = b.get('busbar', '')
    acc, n = [], 1
    def add(name, spec, qty, rem):
        nonlocal n
        acc.append((str(n), name, spec, qty, rem)); n += 1
    if 'SPD' in tags:
        add('Surge Protection Device (SPD)', '8/20µs TYPE 2, In≥40kA, IEC 61643', '1 set', 'Via 6A MCB')
    if 'CT' in tags or ct:
        add('Current Transformer (CT)', ct or '—', 'per phase', '')
    if 'ELR' in tags:
        add('Earth Leakage Relay (ELR)', '—', '1 no.', 'Connected to ZCT')
    if 'ZCT' in tags:
        add('Zero Core Current Transformer (ZCT)', '—', '1 no.', '')
    if 'PM' in tags:
        add('Power Meter (PM)', 'Class 0.5, BMS output', '1 no.', '')
    if 'BMS' in tags:
        add('BMS Interface / Transducer', 'Class 0.5', '1 set', '')
    if busbar:
        add('Insulated CU Busbar', busbar, '1 set', '')
    return acc


# ─── Main ─────────────────────────────────────────────────────────────────────
def build_loadschedule(boards: list, output_path: str) -> str:
    wb = Workbook()

    ws_index = wb.active
    ws_index.title = INDEX_SHEET
    ws_diag = wb.create_sheet(DIAGRAM_SHEET)

    tab_cycle = [DARK_BLUE, MID_BLUE, GREEN, 'ED7D31']
    for i, b in enumerate(boards):
        if not b.get('acc') and (b.get('acc_tags') or b.get('busbar')):
            b['acc'] = _reconstruct_acc(b)
        ws = wb.create_sheet(sheet_name(b.get('id') or f'Board_{i+1}'))
        _build_sheet(ws, b, tab_cycle[i % len(tab_cycle)], with_demand_formula=True)
        # Whole title bar (A1) links back to INDEX
        _link(ws['A1'], INDEX_SHEET)

    _build_index(ws_index, boards)
    _build_diagram(ws_diag, boards)

    wb.save(output_path)
    return output_path


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import json
    if len(sys.argv) < 3:
        print("Usage: python scripts/build_loadschedule.py <boards.json> <out.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding='utf-8') as f:
        data = json.load(f)
    print("Saved:", build_loadschedule(data, sys.argv[2]))
