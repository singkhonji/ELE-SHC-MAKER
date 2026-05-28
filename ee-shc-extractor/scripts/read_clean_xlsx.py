"""
read_clean_xlsx.py — Parse a (possibly user-edited) clean workbook back into
                     board dicts for Stage 2 (Load Schedule).

Round-trips the layout written by build_clean_xlsx.py.  The user may edit the
*values* (fill in descriptions, Connected Load kW, Demand Factor, etc.) but
should keep the labels, the 'CABINET:' markers, the circuit-table columns, and
the blank spacer rows intact.

Usage (when imported):
    read_clean_xlsx(path) -> list[dict]
    Each dict mirrors parse_board() output plus 'acc_tags', 'conn_load', 'demand'.
"""

import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_clean_xlsx import (
    CABINET_MARKER, LBL_SUPPLY, LBL_INC_CB, LBL_INC_CAB,
    LBL_BUSBAR, LBL_ACC, LBL_FLAGS,
)

_LABEL_FIELD = {
    LBL_SUPPLY.lower():   'supply',
    LBL_INC_CB.lower():   'incomer_cb',
    LBL_INC_CAB.lower():  'incomer_cable',
    LBL_BUSBAR.lower():   'busbar',
    LBL_ACC.lower():      '_acc',
    LBL_FLAGS.lower():    '_flags',
}


def _s(v) -> str:
    return '' if v is None else str(v).strip()


def _clean_value(v: str) -> str:
    """Strip placeholder noise so empty/⚠️-only cells become ''."""
    s = _s(v)
    if s in ('—', '-', ''):
        return ''
    # A pure "⚠️ — ... verify" placeholder carries no real data.
    if s.startswith('⚠️') and ('not in DXF' in s or 'refer drawing' in s):
        return ''
    return s


def _parse_acc(summary: str):
    """'SPD | CT 400/5A 5P20,20VA | PM | BMS | ELR | ZCT' -> (tags, ct_ratio)."""
    tags, ct = [], ''
    for tok in _s(summary).split('|'):
        t = tok.strip()
        if not t or t == '—':
            continue
        up = t.upper()
        if up.startswith('CT'):
            tags.append('CT')
            ct = t[2:].strip()
        elif up in ('SPD', 'PM', 'BMS', 'ELR', 'ZCT'):
            tags.append(up)
        else:
            tags.append(t)
    return tags, ct


def _new_board():
    return {
        'id': '', 'supply': '', 'incomer_cb': '', 'incomer_cable': '',
        'busbar': '', 'spd': '', 'ct': '', 'cable_dc': False,
        'acc_tags': [], 'circuits': {},
    }


def read_clean_xlsx(path: str) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Clean Summary'] if 'Clean Summary' in wb.sheetnames else wb.active

    boards = []
    cur = None
    mode = None  # None | 'circuits'

    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)

        # New cabinet block
        if a.startswith(CABINET_MARKER):
            if cur:
                boards.append(cur)
            cur = _new_board()
            cur['id'] = a[len(CABINET_MARKER):].strip()
            mode = None
            continue

        if cur is None:
            continue

        # Circuit-table header → switch to reading circuit rows
        if a == 'Circuit No.':
            mode = 'circuits'
            continue

        if mode == 'circuits':
            if not a:                      # blank spacer = end of this cabinet
                mode = None
                continue
            raw = [ws.cell(r, c).value for c in range(1, 13)]
            row = [_s(v) for v in raw]
            ckt    = row[0]
            desc   = row[1]
            status = (row[10] or '').upper()
            spare  = status == 'SPARE' or desc.upper() == 'SPARE'
            cur['circuits'][ckt] = {
                'desc':      desc,
                'cb_a':      row[2] or '--',
                'cb_brk':    row[5] or '36kA',
                'conn_load': raw[6] if raw[6] is not None else '',   # keep numeric type
                'demand':    raw[7] if raw[7] is not None else '',
                'cable':     (row[8] or '--') if not spare else '--',
                'route':     (row[9] or '--') if not spare else '--',
                'status':    status or ('SPARE' if spare else 'ACTIVE'),
                'spare':     spare,
                'remark':    row[11],
            }
            continue

        # Labelled header row (value in column C / index 3)
        field = _LABEL_FIELD.get(a.rstrip(':').lower() + ':')
        if field is None:
            field = _LABEL_FIELD.get(a.lower())
        if field:
            val = _clean_value(ws.cell(r, 3).value)
            if field == '_acc':
                tags, ct = _parse_acc(ws.cell(r, 3).value)
                cur['acc_tags'] = tags
                cur['ct'] = ct
                cur['spd'] = 'SPD' in tags
            elif field == '_flags':
                cur['cable_dc'] = 'CABLE BY DC MEP' in _s(ws.cell(r, 3).value).upper()
            else:
                cur[field] = val

    if cur:
        boards.append(cur)
    return boards


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    if len(sys.argv) < 2:
        print("Usage: python scripts/read_clean_xlsx.py <clean.xlsx>")
        sys.exit(1)
    for b in read_clean_xlsx(sys.argv[1]):
        print(f"\n{b['id']}  | supply={b['supply']} | incomer={b['incomer_cb']}")
        print(f"   busbar={b['busbar']} | acc={b['acc_tags']} ct={b['ct']} dc={b['cable_dc']}")
        for ck, d in b['circuits'].items():
            print(f"   {ck:7} {d['desc'][:28]:<28} {d['cb_a']:>5} {d['status']:<6} "
                  f"kW={d['conn_load']!r} df={d['demand']!r} {d['cable'][:30]}")
